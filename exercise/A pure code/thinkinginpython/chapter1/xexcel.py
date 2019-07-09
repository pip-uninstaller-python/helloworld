import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.max_row)
for rowofcell in sheet['A1':'E4']:
    # print(rowofcell)
    for cell in rowofcell:
        print(cell.value)
    print('\n')

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countyData = {}

for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)

# print(countyData)


wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}
for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updateproduceSales.xlsx')
